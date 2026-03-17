#!/usr/bin/env python3
"""
Excel Writer for Kosovo VAT Purchase Book (LB.xlsx)
Maps Unstract Vartex Classification JSON output to LB.xlsx structure
"""

import json
import sys
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import time
import fcntl

# Field code to column mapping (based on row 3 of LB.xlsx)
FIELD_CODE_COLUMNS = {
    '[31]': 'G',   # Blerjet dhe importet pa TVSH
    '[32]': 'H',   # Blerjet dhe importet investive pa TVSH
    '[33]': 'I',   # Blerjet dhe importet me TVSH jo të zbritshme
    '[34]': 'J',   # Blerjet dhe importet investive me TVSH jo të zbritshme
    '[35]': 'K',   # Importet (18%)
    '[39]': 'L',   # Importet investive (18%)
    '[43]': 'M',   # Blerjet vendore (18%)
    '[47]': 'N',   # Blerjet investive vendore (18%)
    '[53]': 'O',   # Nota debitore/kreditore (18%)
    '[57]': 'P',   # Fatura e borxhit të keq (18%)
    '[61]': 'Q',   # Rregullimet për të ulur TVSH (18%)
    '[65]': 'R',   # E drejta e kreditimit (Reverse charge)
    # Column S [K1] is calculated - skip
    '[37]': 'T',   # Importet (8%)
    '[41]': 'U',   # Importet investive (8%)
    '[45]': 'V',   # Blerjet vendore (8%)
    '[49]': 'W',   # Blerjet investive vendore (8%)
    '[51]': 'X',   # Blerjet nga fermerët
    '[55]': 'Y',   # Nota debitore/kreditore (8%)
    '[59]': 'Z',   # Fatura e borxhit të keq (8%)
    '[63]': 'AA',  # Rregullimet për të ulur TVSH (8%)
    # Column AB [K2] is calculated - skip
    # Column AC [67] is calculated - skip
    '[68]': 'AD'   # Other
}

def log_message(message, level="INFO"):
    """Log message with timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] [{level}] {message}", file=sys.stderr)

def parse_json_input(json_string):
    """Parse JSON input from Unstract"""
    try:
        data = json.loads(json_string)
        log_message(f"Successfully parsed JSON input")
        return data
    except json.JSONDecodeError as e:
        log_message(f"JSON parsing error: {e}", "ERROR")
        raise

def validate_invoice_data(data):
    """Validate required fields in invoice data"""
    required_fields = ['invoice_data', 'amounts']
    
    for field in required_fields:
        if field not in data:
            raise ValueError(f"Missing required field: {field}")
    
    # Validate invoice_data has required subfields
    invoice_data = data['invoice_data']
    if 'Data' not in invoice_data or not invoice_data['Data']:
        raise ValueError("Missing or empty invoice date (Data)")
    
    log_message("Validation passed")
    return True

def find_first_empty_row(worksheet, start_row=4):
    """Find the first empty row in the worksheet starting from start_row"""
    for row_num in range(start_row, worksheet.max_row + 100):
        # Check if column B (Data) is empty - this indicates an empty row
        if worksheet[f'B{row_num}'].value is None:
            log_message(f"Found empty row at: {row_num}")
            return row_num
    
    # If no empty row found, return max_row + 1
    return worksheet.max_row + 1

def write_invoice_to_excel(excel_path, invoice_json):
    """Write invoice data to LB.xlsx"""
    log_message(f"Opening Excel file: {excel_path}")
    
    # Parse JSON
    data = parse_json_input(invoice_json)
    
    # Validate data
    validate_invoice_data(data)
    
    # Extract data sections
    invoice_data = data['invoice_data']
    amounts = data['amounts']
    vat_classification = data.get('vat_classification', {})
    
    # Open workbook with file locking
    max_retries = 5
    retry_delay = 2
    
    for attempt in range(max_retries):
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Find first empty row
            empty_row = find_first_empty_row(ws)
            log_message(f"Writing to row: {empty_row}")
            
            # Column A: Nr. (row number - can be auto-incremented or left empty)
            # Leave empty for now - user can fill manually or add formula
            
            # Column B: Data (invoice date)
            ws[f'B{empty_row}'] = invoice_data.get('Data', '')
            ws[f'B{empty_row}'].alignment = Alignment(horizontal='left')
            
            # Column C: Numri i faturës (invoice number)
            invoice_number = invoice_data.get("'Numri i faturës", invoice_data.get("Numri i faturës", ''))
            ws[f'C{empty_row}'] = invoice_number
            ws[f'C{empty_row}'].alignment = Alignment(horizontal='left')
            
            # Column D: Emri i shitësit (supplier name)
            ws[f'D{empty_row}'] = invoice_data.get('Emri i shitësit', '')
            ws[f'D{empty_row}'].alignment = Alignment(horizontal='left')
            
            # Column E: Numri Fiskal i shitësit (NUI/NF/NP)
            # Use fiscal number from invoice_data, fallback to supplier_type only if it's a special code (NUI/NF/NP)
            fiscal_number = invoice_data.get('Numri Fiskal i shitësit', '')
            supplier_type = vat_classification.get('supplier_type', '')
            # Only use supplier_type if fiscal_number is empty AND supplier_type is a special code
            if fiscal_number:
                supplier_id = fiscal_number
            elif supplier_type in ['NUI', 'NF', 'NP']:
                supplier_id = supplier_type
            else:
                supplier_id = fiscal_number  # Use empty string if nothing valid
            ws[f'E{empty_row}'] = supplier_id
            ws[f'E{empty_row}'].alignment = Alignment(horizontal='left')
            
            # Column F: Numri i TVSH-së së shitësit (VAT number)
            vat_number = invoice_data.get('Numri i TVSH-së së shitësit', '')
            # If VAT number is null or "null" string, leave empty
            if vat_number and vat_number.lower() != 'null':
                ws[f'F{empty_row}'] = vat_number
                ws[f'F{empty_row}'].alignment = Alignment(horizontal='left')
            
            # Columns G-AD: Amount fields based on field codes
            for field_code, column in FIELD_CODE_COLUMNS.items():
                amount_value = amounts.get(field_code, 0)
                
                # Only write non-zero values to keep Excel clean
                if amount_value and amount_value != 0:
                    ws[f'{column}{empty_row}'] = float(amount_value)
                    ws[f'{column}{empty_row}'].alignment = Alignment(horizontal='right')
                    ws[f'{column}{empty_row}'].number_format = '0.00'
            
            # Copy formulas from row 4 to the new row for calculated columns
            # Column S [K1]: TVSH e zbritshme me 18%
            if ws['S4'].value and isinstance(ws['S4'].value, str) and ws['S4'].value.startswith('='):
                formula = ws['S4'].value
                # Update row reference in formula
                new_formula = formula.replace('4', str(empty_row))
                ws[f'S{empty_row}'] = new_formula
                ws[f'S{empty_row}'].number_format = '0.00'
            
            # Column AB [K2]: TVSH e zbritshme me 8%
            if ws['AB4'].value and isinstance(ws['AB4'].value, str) and ws['AB4'].value.startswith('='):
                formula = ws['AB4'].value
                new_formula = formula.replace('4', str(empty_row))
                ws[f'AB{empty_row}'] = new_formula
                ws[f'AB{empty_row}'].number_format = '0.00'
            
            # Column AC [67]: Totali i TVSH-së
            if ws['AC4'].value and isinstance(ws['AC4'].value, str) and ws['AC4'].value.startswith('='):
                formula = ws['AC4'].value
                new_formula = formula.replace('4', str(empty_row))
                ws[f'AC{empty_row}'] = new_formula
                ws[f'AC{empty_row}'].number_format = '0.00'
            
            # Save workbook
            wb.save(excel_path)
            wb.close()
            
            log_message(f"Successfully wrote invoice to row {empty_row}")
            
            # Return success with row number
            result = {
                "status": "success",
                "row_number": empty_row,
                "invoice_number": invoice_number,
                "supplier": invoice_data.get('Emri i shitësit', ''),
                "date": invoice_data.get('Data', '')
            }
            
            print(json.dumps(result))
            return 0
            
        except PermissionError:
            if attempt < max_retries - 1:
                log_message(f"File locked, retrying in {retry_delay}s... (attempt {attempt + 1}/{max_retries})", "WARNING")
                time.sleep(retry_delay)
            else:
                log_message("Failed to access file after multiple retries", "ERROR")
                raise
        except Exception as e:
            log_message(f"Error writing to Excel: {e}", "ERROR")
            raise

def main():
    """Main execution function - accepts JSON from stdin or command line"""
    if len(sys.argv) < 2:
        print(json.dumps({
            "status": "error",
            "message": "Usage: python excel_writer_lb.py <excel_path> [json_data or pipe from stdin]"
        }))
        return 1
    
    excel_path = sys.argv[1]
    
    # Read JSON data from stdin or command line argument
    if len(sys.argv) >= 3:
        # JSON provided as command line argument
        json_data = sys.argv[2]
        log_message("Reading JSON from command line argument")
    else:
        # Read from stdin
        log_message("Reading JSON from stdin")
        json_data = sys.stdin.read().strip()
        if not json_data:
            print(json.dumps({
                "status": "error",
                "message": "No JSON data received from stdin"
            }))
            return 1
    
    try:
        # Validate Excel file exists
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        # Write to Excel
        return write_invoice_to_excel(excel_path, json_data)
        
    except Exception as e:
        log_message(f"Fatal error: {e}", "ERROR")
        print(json.dumps({
            "status": "error",
            "message": str(e)
        }))
        return 1

if __name__ == "__main__":
    sys.exit(main())
